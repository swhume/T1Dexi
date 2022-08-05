import argparse
import os
from openpyxl import load_workbook
import xlsxwriter as XLS
import requests

"""
map2datasets.py extracts the variables from the mapping spreadsheet and generates a Define-XML v2.1 metadata worksheet
for datasets.
Example Cmd-line (optional args):
    python map2datasets -i ./path/to/mapping_spec.xlsx -o ./path/to/datasets_ws.xlsx
"""

# odmlib worksheet column headers to variables
header = ["OID", "Dataset", "Description", "Class", "Structure", "Purpose", "Repeating",
          "Reference Data",	"Comment", "IsNonStandard", "StandardOID", "HasNoData"]

library_api_key = "e9a7d1b9bf1a4036ae7b25533a081565"

# SDTM mapping spreadsheet worksheets to skip as they do not contain variables for a given domain
worksheet_skip = ["T1Dexi SDTM Summary", "T1Dexi Tables", "Domains", "Sheet1"]
# dataset ordering for SDTM define-xml based on MSG 2.0
class_order = ["TRIAL DESIGN", "SPECIAL PURPOSE", "INTERVENTIONS", "EVENTS", "FINDINGS", "FINDINGS ABOUT",
               "RELATIONSHIP", "STUDY REFERENCE"]

# name and path of the input SDTM mapping spreadsheet and default -i CLI arg value - assumes child data dir
# excel_map_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'SDTM-mapping-spec-02Feb2022.xlsx')
excel_map_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'SDTM-mapping-spec-20220406.xlsx')
# name and path of the output odmlib variables spreadsheet and default -o CLI arg value - assumes child data dir
excel_define_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'datasets_test.xlsx')


def write_header_row(worksheet, header_format):
    """
    write the worksheet header column labels to the odmlib worksheet
    :param worksheet: the odmlib worksheet to write the column headers to
    :param header_format: the header format to indicate the style of the header columns
    """
    for c, header_label in enumerate(header):
        worksheet.write(0, c, header_label, header_format)


def create_define_sheet(table, sheet_name, workbook, header_format):
    """
    write variable definition rows to the odmlib worksheet
    :param table: list of row dictionaries with content to write to the worksheet
    :param sheet_name: name of the odmlib worksheet to create to write variable content
    :param workbook: workbook object to add the worksheet to
    :param header_format: format for the header_row used when writing the column headers
    """
    worksheet = workbook.add_worksheet(sheet_name)
    write_header_row(worksheet, header_format)
    for r, row in enumerate(table):
        for c, col_name in enumerate(header):
            try:
                worksheet.write(r+1, c, row[col_name])
            except Exception as ex:
                print(ex)


def get_endpoint_from_library(endpoint, api_key):
    """
    retrieve a codelist from the CDISC Library using the provided endpoint
    :param endpoint: endpoint string used to create the API call to the Library to retrieve a codelist
    :param api_key: string CDISC Library API key
    :return: json results from Library GET request for a specified codelist
    """
    base_url = "https://library.cdisc.org/api"
    headers = {"Accept": "application/json", "User-Agent": "crawler", "api-key": api_key}
    r = requests.get(base_url + endpoint, headers=headers)
    if r.status_code == 200:
        return r.json()
    else:
        print(f"HTTPError {r.status_code} for url {base_url + endpoint}")


def load_class_names(api_key):
    """
    lookup the class names for the datasets in SDTMIG 3.3
    :param api_key: string; API key needed to authenticate and access the CDISC Library API
    :return: dictionary of class names associated with the dataset
    """
    endpoint = "/mdr/sdtmig/3-3/classes"
    classes = get_endpoint_from_library(endpoint, api_key)
    class_sub_vals = lookup_class_sub_values(api_key)
    class_names = {}
    for cl_dict in classes["_links"]["classes"]:
        class_dict = get_endpoint_from_library(cl_dict["href"], api_key)
        class_names[class_dict["label"]] = find_class_match(class_dict["name"], class_sub_vals)
    return class_names


def find_class_match(name, sub_vals):
    """
    find the Define-XML CT term based on the class name for the dataset
    :param name: string; name of the class associated with a dataset in the Library
    :param sub_vals: dictionary of Define-XML class terms (submission values) to look-up by name
    :return: string; return the Define-XML Class term
    """
    for sv, synonyms in sub_vals.items():
        if name in synonyms or name.upper() in synonyms or name.upper() == sv:
            return sv
        elif name == "General Observations":
            return name.upper()
    return name


def lookup_class_sub_values(api_key):
    """
    lookup Define-XML class controlled terminology
    :param api_key: string; API key to access the Library API
    :return: dictionary with the synonyms needed to look-up the Define-XML CT Class term
    """
    endpoint = "/mdr/ct/packages/define-xmlct-2021-12-17/codelists/C103329"
    cl = get_endpoint_from_library(endpoint, api_key)
    class_sub_vals = {}
    for term in cl["terms"]:
        class_sub_vals[term["submissionValue"]] = term["synonyms"]
    return class_sub_vals


def lookup_domain_endpoint(domain):
    """
    given the domain short name build the endpoint to retrieve the dataset metadata from the Library
    :param domain: string; short domain name
    :return: string; endpoint URL
    """
    if domain in ["DX", "DI"]:
        endpoint = "/mdr/sdtmig/md-1-1/datasets/" + domain
    elif "SUPP" in domain:
        endpoint = "/mdr/sdtmig/3-3/datasets/SUPPQUAL"
    elif "FA" in domain:
        endpoint = "/mdr/sdtmig/3-3/datasets/FA"
    else:
        endpoint = "/mdr/sdtmig/3-3/datasets/" + domain
    return endpoint


def process_map_sheet(sheet, def_workbook, header_format, api_key):
    """
    process the content in the SDTM mapping spreadsheet to generate the define-xml metadata worksheet
    :param sheet: SDTM mapping spreadsheet worksheet to parse
    :param domain: 2-letter domain abbreviation used to identify the domain for a variable
    :param def_workbook: obmlib define-xml workbook to add variables worksheet to
    :param header_format: format for worksheet column headers
    """
    class_names = load_class_names(api_key)
    domain = sheet["A3"].value
    rows = []
    row_nbr = 3
    while domain:
        row_dict = {key: "" for key in header}
        endpoint = lookup_domain_endpoint(domain)
        library = get_endpoint_from_library(endpoint, api_key)
        row_dict["OID"] = "IG." + domain
        row_dict["Dataset"] = domain
        row_dict["Description"] = sheet["B" + str(row_nbr)].value
        if library is None:
            print()
        row_dict["Class"] = class_names[library["_links"]["parentClass"]["title"]]
        row_dict["Structure"] = library["datasetStructure"]
        row_dict["Purpose"] = "Tabulation"
        if domain == "DM":
            row_dict["Repeating"] = "No"
        else:
            row_dict["Repeating"] = "Yes"
        row_dict["Reference Data"] = "No"
        row_dict["Comment"] = ""
        row_dict["IsNonStandard"] = ""
        row_dict["StandardOID"] = ""
        row_dict["HasNoData"] = ""
        rows.append(row_dict)
        row_nbr += 1
        # make sure all domains are listed in the Domains Used table as blank cells will break out of the loop
        domain = sheet["A" + str(row_nbr)].value
    ordered_rows = sort_domain_order(rows)
    create_define_sheet(ordered_rows, sheet.title, def_workbook, header_format)


def sort_domain_order(rows):
    sorted_table = []
    for class_name in class_order:
        row_subset = []
        for row in rows:
            if row["Class"] == class_name:
                row_subset.append(row)
        sorted_rows = sort_by_dataset_name(row_subset)
        sorted_table.extend(sorted_rows)
    return sorted_table


def sort_by_dataset_name(rows):
    sorted_table = []
    dataset_names = [row["Dataset"] for row in rows]
    dataset_names.sort()
    for name in dataset_names:
        for row in rows:
            if row["Dataset"] == name:
                sorted_table.append(row)
    return sorted_table


def set_cmd_line_args():
    """
    get the command-line arguments needed to convert the Excel input file into Define-XML
    :return: return the argparse object with the command-line parameters
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input_file", help="path and file name of the SDTM map spreadsheet file",
                        required=False, dest="map_xls", default=excel_map_file)
    parser.add_argument("-o", "--output_file", help="path and file name of Define-XML v2.1 metadata spreadsheet file",
                        required=False, dest="define_xls", default=excel_define_file)
    parser.add_argument("-a", "--api_key", help="CDISC Library API key",
                        required=False, dest="api_key", default=library_api_key)
    args = parser.parse_args()
    return args


def main():
    """
    main driver application that processes the SDTM mapping spreadsheet and creates and odmlib dataset worksheet
    """
    args = set_cmd_line_args()
    def_workbook = XLS.Workbook(args.define_xls, {"strings_to_numbers": False})
    header_format = def_workbook.add_format({"bold": True, "bg_color": "#CCFFFF", "border": True, "border_color": "black"})
    map_workbook = load_workbook(filename=args.map_xls, read_only=False, data_only=True)
    sheet = map_workbook["Domains"]
    process_map_sheet(sheet, def_workbook, header_format, args.api_key)
    def_workbook.close()


if __name__ == '__main__':
    main()
