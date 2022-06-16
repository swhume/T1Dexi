from openpyxl import load_workbook
import xlsxwriter as XLS
import os.path
import json
import argparse

"""
map2subsets.py generates content for the ValueLevel, WhereClauses, and CodeLists worksheets in the T1Dexi odmlib
metadata spreadsheet. The output spreadsheet produced includes the previously mentioned worksheets and this content
can be pasted into the corresponding metadata worksheets. Content is pulled from the SDTM mapping spreadsheet, but the
mapping spreadsheet doesn't contain all the necessary information in a way that can be processed in a straightforward 
manner. The content is found in the CODELIST - VARIABLE NAME columns after the variable name columns.
The codelists dictionary listed below adds information needed to work with content from the mapping spreadsheet. 
If there are changes to the study that impact value level metadata this dictionary may need to
be updated.

Example Cmd-line (optional args):
    python map2variables -i ./path/to/mapping_spec.xlsx -o ./path/to/define-worksheets.xlsx
"""


# --- since splitting on ", " could also get rid of the space in the Masters codelist
# codelists are those codelists documented in columns after the variables (CODELIST - VARIABLE NAME)
# TODO document dictionary structure
codelists = {"DM.RACE": {"IsNonStandard": ["C74457"], "VLM": "No", "type": ["text"], "whereclause": []},
             "DM.ETHNICITY": {"IsNonStandard": ["C66790"], "VLM": "No", "type": ["text"], "whereclause": []},
             "DX.DXTRT": {"IsNonStandard": ["Yes"], "VLM": "No", "type": ["text"], "whereclause": []},
             "FA.FAORRES": {"IsNonStandard": ["Yes", "Yes", "Yes", "Yes", "Yes"], "VLM": "Yes",
                            "type": ["text", "text", "text", "text", "text"],
                            "whereclause": [{"variable": "FATESTCD", "comparator": "EQ", "value": "AGE"},
                                            {"variable": "FATESTCD", "comparator": "EQ", "value": "SICKTODY"},
                                            {"variable": "FATESTCD", "comparator": "EQ", "value": "INSCHFL"},
                                            {"variable": "FATESTCD", "comparator": "EQ", "value": "STRESTDY"},
                                            {"variable": "FATESTCD", "comparator": "EQ", "value": "SLEEPQLT"}
                                        ]
                            },
             "FADX.FAORRES": {"IsNonStandard": ["Yes", "Yes", "No"], "VLM": "Yes", "type": ["text", "text", "integer"],
                              "whereclause": [{"variable": "FAOBJ", "comparator": "EQ", "value": "INSULIN PUMP OR CLOSED LOOP"},
                                            {"variable": "FAOBJ", "comparator": "EQ", "value": "CGM"},
                                            {"variable": "FAOBJ", "comparator": "EQ", "value": "CGM USE LAST MONTH"}
                                        ]
                              },
             "FADX.DISINSEX": {"IsNonStandard": ["Yes"], "VLM": "No", "type": ["text"], "whereclause": []},
             "FADX.SUSPINS": {"IsNonStandard": ["Yes"], "VLM": "No", "type": ["text"], "whereclause": []},
             "LB.LBTMINT": {"IsNonStandard": ["Yes"], "VLM": "No", "type": ["text"], "whereclause": []},
             "SC.SCORRES": {"IsNonStandard": ["Yes", "Yes"], "VLM": "Yes", "type": ["text", "text"],
                            "whereclause": [{"variable": "SCTESTCD", "comparator": "EQ", "value": "EDULEVEL"},
                                            {"variable": "SCTESTCD", "comparator": "EQ", "value": "INCMLVL"}
                                            ]
                            },
             "QS.TESTCD": {"IsNonStandard": ["C141665", "Yes", "Yes"], "VLM": "Yes", "type": ["text", "text", "text"],
                           "whereclause": [{"variable": "QSCAT", "comparator": "EQ", "value": "IPAQ SHORT - SELF ADMINISTERED VERSION"},
                                           {"variable": "QSCAT", "comparator": "EQ", "value": "CLARK HYPOGLYCEMIA UNAWARENESS SURVEY SELF ADMINISTERED"},
                                           {"variable": "QSCAT", "comparator": "EQ", "value": "PITTSBURG SLEEP QUALITY INDEX"}
                           ],
                           "subset_terms": ["IPA0401, IPA0402, IPA0403, IPA0404, IPA0405, IPA0406, IPA0407",
                                            "CHU01, CHU02, CHU03, CHU04, CHU05, CHU06, CHU07, CHU08",
                                            "PSQI01, PSQI02, PSQI03, PSQI04, PSQI05A, PSQI05B, PSQI05C, PSQI05D, PSQI05E, PSQI05F, PSQI05G, PSQI05H, PSQI05I, PSQI05KJA, PSQI05JB, PSQI06, PSQI07, PSQI08, PSQI09"]
                           },
             "QS.TEST": {"IsNonStandard": ["C141664", "Yes", "Yes"], "VLM": "Yes", "type": ["text", "text", "text"],
                           "whereclause": [{"variable": "QSCAT", "comparator": "EQ", "value": "IPAQ SHORT - SELF ADMINISTERED VERSION"},
                                           {"variable": "QSCAT", "comparator": "EQ", "value": "CLARK HYPOGLYCEMIA UNAWARENESS SURVEY SELF ADMINISTERED"},
                                           {"variable": "QSCAT", "comparator": "EQ",  "value": "PITTSBURG SLEEP QUALITY INDEX"}
                                           ],
                           "subset_terms": ["IPA04-Days Vigorous Physical Activities, IPA04-Vigorous Physical Hr & Min per Day, \
IPA04-Days Moderate Physical Activities, IPA04-Moderate Physical Hr & Min per Day, IPA04-Days Walk at Least 10 Minutes, \
IPA04-Walking Hr & Min per Day, IPA04-Sitting Weekday Hr & Min per Day", "CHU01-Category Best Describes You, \
CHU01-Lost Some Symptoms Blood Sugar Low, CHU01-Moderate Hypo Espisodes-P6M, CHU01-Severe Hypo Episodes-P1Y, \
CHU01-<70 mg/dL w/ Symptoms-P1M, CHU01-<70 mg/dL w/o  Symptoms-P1M, CHU01-How Low Blood Sug Before Symptoms, \
CHU01-Extent Can Tell Blood Sugar is Low", "PSPSQI01-Time Usually Go To Bed-P1M, PSQI01-Minutes Taken to Fall Asleep-P1M, \
PSQI01-Time Usually Wake Up-P1M, PSQI01-Hours Actual Sleep at Night-P1M, PSPSQI01-Can't Get to Sleep in 30 Mins-P1M, \
PSQI01-Wake up Mid Night Early Morn-P1M, PSQI01-Have to Use Bathroom-P1M, PSQI01-Cannot Breathe  Comfortably-P1M, \
PSQI01-Cough or Snore Loudly-P1M, PSQI01-Feel Too Cold-P1M, PSQI01-Feel Too Hot-P1M, PSQI01-Have Bad Dreams-P1M, \
PSQI01-Have Pain-P1M, PSQI01-Trouble Sleep Other reason(s)-P1M, PSQI01-Describe Other reason(s)-P1M, \
PSQI01-Sleep Medication-P1M, PSQI01-Trouble Staying Awake Social-P1M, PSQI01-Problem Keeping Up Enthusiasm-P1M, \
PSQI01-Rate Overall Sleep Quality-P1M"]
                           }
             }

worksheet_skip = ["T1Dexi SDTM Summary", "T1Dexi Tables", "Domains", "Sheet1"]
excel_map_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'SDTM-mapping-spec-20220406.xlsx')
subset_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'cl_subsets.json')
excel_define_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'codelist_subsets-test.xlsx')

# worksheet headers
cl_header = ["OID", "Name", "NCI Codelist Code", "Data Type", "Order", "Term", "NCI Term Code", "Decoded Value",
          "Comment", "IsNonStandard", "StandardOID"]
wc_header = ["OID", "Dataset", "Variable", "Comparator", "Value", "Comment"]
vlm_header = ["OID", "Order", "Dataset", "Variable", "ItemOID", "Where Clause", "Data Type", "Length",
              "Significant Digits", "Format", "Mandatory", "Codelist", "Origin Type", "Origin Source", "Pages",
              "Method", "Predecessor", "Comment"]


def write_subset_file():
    """
    write the codelist subsets dictionary to file as JSON
    """
    with open(subset_file, 'w') as file_out:
        json.dump(codelists, file_out)


def process_map_sheet(sheet, domain):
    """
    process the content in the SDTM mapping spreadsheet for content to use in the define-xml metadata worksheets
    :param sheet: SDTM mapping spreadsheet worksheet to parse
    :param domain: 2-letter domain abbreviation used to identify the domain for a variable
    """
    # assumes not more than 6 VLM (max_row=17)
    for col_num, col in enumerate(sheet.iter_cols(min_col=2,max_col=sheet.max_column, min_row=1, max_row=17, values_only=True)):
        # row_dict = {key: "" for key in header}
        if col and "CODELIST" in str(col[0]):
            print(f"processing codelist {col[0]}...")
            codelist, var_name = col[0].split(" - ")
            subset_codes = []
            for row_num, cell in enumerate(col):
                if row_num > 9 and cell:
                    subset_codes.append(cell)
                    # if more than one subset then have VLM and must lookup testcds
            print(f"found {len(subset_codes)} subsets for {var_name} in domain {domain}")
            codelists[domain + "." + var_name]["domain"] = domain
            if subset_codes:
                codelists[domain + "." + var_name]["subset_terms"] = subset_codes


def process_codelists():
    """
    process the codelists dictionary that has the SDTM mapping spreadsheet content added to create codelist XLS rows
    :return: list of dictionaries with codelist worksheet rows
    """
    rows = []
    for key, cl in codelists.items():
        domain, variable = key.split(".")
        for idx, data_type in enumerate(cl["type"]):
            row = {key: "" for key in cl_header}
            if cl["IsNonStandard"][idx] == "No":
                continue
            if cl["VLM"] == "No":
                row["OID"] = "CL." + domain + "." + variable
                row["Name"] = "Codelist for " + domain + " " + variable
            else:
                wc_value = generate_wc_name(cl["whereclause"][idx]["value"])
                row["OID"] = "CL." + domain + "." + variable + "." + wc_value
                row["Name"] = "Codelist for " + domain + " " + variable + " where " + wc_value
            if len(cl["IsNonStandard"][idx]) > 3:
                row["NCI Codelist Code"] = cl["IsNonStandard"][idx]
            else:
                row["NCI Codelist Code"] = ""
            row["Data Type"] = cl["type"][idx]
            order_num = 1
            for term in cl["subset_terms"][idx].split(", "):
                row["Order"] = order_num
                term_row = row.copy()
                term_row["Term"] = term
                term_row["NCI Term Code"] = ""
                term_row["Decoded Value"] = ""
                term_row["Comment"] = ""
                if cl["IsNonStandard"][idx] == "Yes":
                    term_row["IsNonStandard"] = "Yes"
                    term_row["StandardOID"] = ""
                else:
                    term_row["IsNonStandard"] = ""
                    term_row["StandardOID"] = "STD.2"
                rows.append(term_row)
                order_num += 1
    return rows


def find_cl_max_length(codelist):
    """
    find and return the length of the longest term in the codelist
    :param codelist: codelist of terms
    :return: integer length of the longest codelist item
    """
    max_length = 0
    for term in codelist.split(", "):
        if len(term) > max_length:
            max_length = len(term)
    return max_length


def find_number_length(domain, variable):
    """
    placeholder to determine the length of numeric fields assigned as value level metadata
    :param domain: the domain / dataset in which the variable exists
    :param variable: the name of the variable
    :return: integer indicating number length - UPDATE THIS DEFAULT VALUE TO GET ACTUAL LENGTH
    """
    return 3


def find_number_sigdigits(domain, variable):
    """
    placeholder to determine the number of significant digits for numeric fields assigned as value level metadata
    :param domain: the domain / dataset in which the variable exists
    :param variable: the name of the variable
    :return: integer indicating number of significant digits - UPDATE THIS DEFAULT TO GET ACTUAL SIGNIFICANT DIGITS
    """
    return 2


def process_where_clauses():
    """
    process the codelist dictionary with SDTM mapping content to generate the Where Clause metadata XLS rows
    :return: list of dictionaries with where clause worksheet rows
    """
    rows = []
    for key, cl in codelists.items():
        domain, variable = key.split(".")
        for idx, wc in enumerate(cl["whereclause"]):
            row = {key: "" for key in wc_header}
            if cl["VLM"] == "No":
                continue
            wc_value = generate_wc_name(wc["value"])
            row["OID"] = "WC." + domain + "." + variable + "." + wc_value
            row["Dataset"] = domain
            row["Variable"] = wc["variable"]
            row["Comparator"] = wc["comparator"]
            row["Value"] = wc["value"]
            row["Comment"] = ""
            rows.append(row)
    return rows


def process_vlm():
    """
    process the codelist dictionary with SDTM mapping content to generate the Value Level metadata XLS rows
    :return: list of dictionaries with value level metadata worksheet rows
    """
    rows = []
    for key, cl in codelists.items():
        domain, variable = key.split(".")
        for idx, wc in enumerate(cl["whereclause"]):
            row = {key: "" for key in wc_header}
            if cl["VLM"] == "No":
                continue
            wc_value = generate_wc_name(wc["value"])
            row["Where Clause"] = "WC." + domain + "." + variable + "." + wc_value
            row["ItemOID"] = "IT." + domain + "." + variable + "." + wc_value
            row["OID"] = "VL." + domain + "." + variable
            row["Dataset"] = domain
            row["Variable"] = variable
            row["Data Type"] = cl["type"][idx]
            if row["Data Type"] == "text":
                row["Codelist"] = "CL." + domain + "." + variable + "." + wc_value
                row["Length"] = find_cl_max_length(row["Codelist"])
                row["Significant Digits"] = ""
                row["Format"] = ""
            elif row["Data Type"] == "integer":
                row["Codelist"] = ""
                row["Length"] = find_number_length(domain, variable)
                row["Significant Digits"] = ""
                row["Format"] = ""
            else:
                row["Codelist"] = ""
                row["Length"] = find_number_length(domain, variable)
                row["Significant Digits"] = find_number_sigdigits(domain, variable)
                row["Format"] = str(find_number_length(domain, variable)) + "." + str(row["Significant Digits"])
            # how determine mandatory for VLM?
            row["Mandatory"] = "No"
            row["Order"] = idx + 1
            row["Origin Type"] = ""
            row["Origin Source"] = ""
            row["Pages"] = ""
            row["Method"] = ""
            row["Predecessor"] = ""
            row["Comment"] = ""
            rows.append(row)
    return rows


def generate_wc_name(wc_name):
    wc_name_nodash = wc_name.replace("-", " ")
    wc_name_nodash = " ".join(wc_name_nodash.split())
    wc_name_value = wc_name_nodash.replace(" ", "-")
    return wc_name_value


def write_codelist_to_xls(worksheet, rows, header, row_nbr=0):
    """
    write code list, value level, or where clause rows to the odmlib worksheet
    :param worksheet: odmlib worksheet object to write to
    :param rows: a list of dictionaries (rows) to write to the worksheet
    :param row_nbr: integer row number that indicates which worksheet row to begin appending terms
    :return: integer row number that indicates the next row to begin appending content
    """
    for row in rows:
        row_nbr += 1
        for c, col_name in enumerate(header):
            worksheet.write(row_nbr, c, row[col_name])
    return row_nbr


def write_header_row(worksheet, header, header_format):
    """
    write the worksheet header column labels to a given odmlib worksheet
    write the header row for a given worksheet
    :param worksheet: the odmlib worksheet to write the column headers to
    :param header: the header for the worksheet to be generated
    :param header_format: the header format to indicate the style of the header columns
    """
    for c, header_label in enumerate(header):
        worksheet.write(0, c, header_label, header_format)


def set_cmd_line_args():
    """
    get the command-line arguments needed to convert the Excel input file into Define-XML
    :return: return the argparse object with the command-line parameters
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input_file", help="path and file name of the SDTM map spreadsheet file",
                        required=False, dest="map_xls", default=excel_map_file)
    parser.add_argument("-o", "--output_file", help="path and file name of Define-XML v2.1 metadata spreadsheet file",
                        required=False, dest="define_xls", default=excel_define_file)
    args = parser.parse_args()
    return args


def main():
    """
    main driver for creating the odmlib worksheet for ValueLists, WhereClauses, and CodeLists
    """
    args = set_cmd_line_args()
    def_workbook = XLS.Workbook(args.define_xls, {"strings_to_numbers": False})
    header_format = def_workbook.add_format({"bold": True, "bg_color": "#CCFFFF", "border": True, "border_color": "black"})
    map_workbook = load_workbook(filename=args.map_xls, read_only=False, data_only=True)
    for sheet in map_workbook.worksheets:
        if sheet.title not in worksheet_skip:
            print(f"processing {sheet.title}...")
            domain = sheet.title.split()
            process_map_sheet(sheet, domain[0])
    rows = process_codelists()
    write_subset_file()

    # create codelist subset worksheet
    worksheet = def_workbook.add_worksheet("codelists")
    write_header_row(worksheet, cl_header, header_format)
    write_codelist_to_xls(worksheet, rows, cl_header)

    # create where clause worksheet
    rows = process_where_clauses()
    wc_worksheet = def_workbook.add_worksheet("whereclauses")
    write_header_row(wc_worksheet, wc_header, header_format)
    write_codelist_to_xls(wc_worksheet, rows, wc_header)

    # create VLM
    rows = process_vlm()
    vlm_worksheet = def_workbook.add_worksheet("valuelevel")
    write_header_row(vlm_worksheet, vlm_header, header_format)
    write_codelist_to_xls(vlm_worksheet, rows, vlm_header)

    def_workbook.close()


if __name__ == '__main__':
    main()
