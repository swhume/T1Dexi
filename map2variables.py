from openpyxl import Workbook, load_workbook
import xlsxwriter as XLS
import argparse
import os.path
import re

"""
map2variables.py extracts the variables from the mapping spreadsheet and generates a Define-XML v2.1 metadata worksheet
for variables.
Example Cmd-line (optional args):
    python map2variables -i ./path/to/mapping_spec.xlsx -o ./path/to/variables_ws.xlsx
"""

# odmlib worksheet column headers to variables
header = ["OID", "Order", "Dataset", "Variable", "Label", "Data Type", "Length", "Significant Digits",
          "Format",	"KeySequence", "Mandatory", "CodeList", "Valuelist", "Origin Type", "Origin Source",
          "Pages", "Method", "Predecessor", "Role", "Comment", "IsNonStandard", "HasNoData"]

# class names defined to format SDTM mapping spreadsheet content for use in odmlib and define-xml
format_class = ["Name", "Label", "Type", "Codelist", "Role", "Notes", "Core", "DataType", "Length", "SignificantDigits"]

# SDTM mapping spreadsheet worksheets to skip as they do not contain variables for a given domain
worksheet_skip = ["T1Dexi SDTM Summary", "T1Dexi Tables", "Domains", "Sheet1"]

# for variables like STUDYID and USUBJID - just define these variables once
common_variables = ["STUDYID", "USUBJID", "SPDEVID"]

# TODO FAML will have separarte datasets for meal, daily, item
# key sequences for datasets
key_sequence = {
    "CM": ["STUDYID", "USUBJID", "CMTRT", "CMSTDTC"],
    "DI": ["STUDYID", "SPDEVID", "DISEQ", "DIPARMCD"],
    "DM": ["STUDYID", "USUBJID"],
    "DX": ["STUDYID", "USUBJID", "DXTRT", "DXDTC"],
    "FA": ["STUDYID", "USUBJID", "FATESTCD", "FAOBJ", "FADTC"],
    "FACM": ["STUDYID", "USUBJID", "FATESTCD", "FAOBJ", "FADTC"],
    "FADX": ["STUDYID", "USUBJID", "FATESTCD", "FAOBJ", "FADTC"],
    "FAML": ["STUDYID", "USUBJID", "FATESTCD", "FAOBJ", "FADTC"],
    "FALB": ["STUDYID", "USUBJID", "FATESTCD", "FAOBJ", "FADTC"],
    "FAPR": ["STUDYID", "USUBJID", "FATESTCD", "FAOBJ", "FADTC"],
    "LB": ["STUDYID", "USUBJID", "LBTESTCD", "LBDTC"],
    "ML": ["STUDYID", "USUBJID", "MLTRT", "MLSTDTC"],
    "NV": ["STUDYID", "USUBJID", "NVTESTCD", "NVDTC"],
    "RELREC": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "RELID"],
    "PR": ["STUDYID", "USUBJID", "PRTRT", "PRSTDTC"],
    "RP": ["STUDYID", "USUBJID", "RPTESTCD", "RPDTC"],
    "SUPPDM": ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL", "QNAM"],
    "QS": ["STUDYID", "USUBJID", "QSTESTCD", "QSDTC"],
    "SC": ["STUDYID", "USUBJID", "SCTESTCD", "SCDTC"],
    "VS": ["STUDYID", "USUBJID", "VSTESTCD", "VSDTC"]
}

# name and path of the input SDTM mapping spreadsheet and default -i CLI arg value - assumes child data dir
excel_map_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'SDTM-mapping-spec-02Feb2022.xlsx')
# name and path of the output odmlib variables spreadsheet and default -o CLI arg value - assumes child data dir
excel_define_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'variables_test.xlsx')


class Name:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: the 2-letter domain abbreviation that the variable belongs to
        :param col_num: integer value used to set the order number
        """
        if cell is None:
            cell = ""
        if cell in common_variables:
            row_dict["OID"] = "IT." + cell
        else:
            row_dict["OID"] = "IT." + domain + "." + cell
        row_dict["Variable"] = cell
        row_dict["Dataset"] = domain
        row_dict["Order"] = col_num + 1
        if domain in key_sequence and cell in key_sequence[domain]:
            row_dict["KeySequence"] = key_sequence[domain].index(cell) + 1


class Label:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: the 2-letter domain abbreviation that the variable belongs to
        :param col_num: integer value used to set the order number
        """
        row_dict["Label"] = cell


class Type:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: the 2-letter domain abbreviation that the variable belongs to
        :param col_num: integer value used to set the order number
        """
        if cell == "Num":
            row_dict["Data Type"] = "integer"
        else:
            row_dict["Data Type"] = "text"


class Codelist:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: the 2-letter domain abbreviation that the variable belongs to
        :param col_num: integer value used to set the order number
        """
        if not cell:
            row_dict["Codelist"] = ""
        elif len(str(cell)) == 2:
            # a domain value
            row_dict["CodeList"] = "CL.DOMAIN." + cell
        elif "ISO 8601" in str(cell):
            # an ISO 8601 date format
            row_dict["Comment"] = "COM.ISO8601"
        elif "ISO 3166-1" in str(cell):
            row_dict["Comment"] = "COM.ISO3166-1"
        elif re.search("^C\d+", str(cell)):
            row_dict["CodeList"] = "CL." + cell
        elif "Non-Standard Variable (NSV)" in str(cell):
            row_dict["IsNonStandard"] = "Yes"
        else:
            print(f"Unknown value in codelist cell: {cell} for column: {col_num}")


class Role:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: not used
        :param col_num: not used
        """
        row_dict["Role"] = cell


class Notes:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        # not using notes at this time
        pass


class Core:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: not used
        :param col_num: not used
        """
        if cell and "Req" in cell:
            row_dict["Mandatory"] = "Yes"
        else:
            row_dict["Mandatory"] = "No"


class DataType:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: not used
        :param col_num: not used
        """
        if cell == "varchar":
            row_dict["Data Type"] = "text"
        elif cell == "int":
            row_dict["Data Type"] = "integer"
        elif cell == "datetime":
            row_dict["Data Type"] = "datetime"
        elif cell == "bit":
            row_dict["Data Type"] = "text"
            row_dict["Length"] = 1
        else:
            row_dict["Data Type"] = "text"


class Length:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: not used
        :param col_num: not used
        """
        if row_dict["Data Type"] == "text":
            if cell:
                row_dict["Length"] = cell
            else:
                # sets default length value to 200 if none provided
                if not row_dict["Length"]:
                    row_dict["Length"] = 200
        elif row_dict["Data Type"] == "integer":
            if cell:
                row_dict["Length"] = cell
            else:
                # sets default length value to 4 if none provided
                row_dict["Length"] = 4
        elif row_dict["Data Type"] == "datetime":
            row_dict["Length"] = ""
        elif row_dict["Data Type"] == "float":
            if cell:
                if "." in str(cell):
                    row_dict["Length"], row_dict["Significant Digits"] = str(cell).split(".")
                else:
                    row_dict["Length"] = cell
            else:
                # sets default length value to 8 if none provided
                row_dict["Length"] = 8


class SignificantDigits:
    def format_content(self, cell, row_dict, domain="", col_num=0):
        """
        formats the variable name into column content for the odmlib Define-XML variables worksheet
        :param cell: SDTM mapping spreadsheet cell value to convert to odmlib Define-XML
        :param row_dict: dictionary for the odmlib Define-XML output format
        :param domain: not used
        :param col_num: not used
        """
        if row_dict["Data Type"] == "float":
            if cell:
                row_dict["Significant Digits"] = cell
            else:
                row_dict["Significant Digits"] = ""


def write_header_row(worksheet, header_format):
    """
    write the worksheet header column labels to the odmlib worksheet
    :param worksheet: the odmlib worksheet to write the column headers to
    :param header_format: the header format to indicate the style of the header columns
    """
    for c, header_label in enumerate(header):
        worksheet.write(0, c, header_label, header_format)


def create_define_sheet(table, sheet_name, workbook, header_format):
    """
    write variable definition rows to the odmlib worksheet
    :param table: list of row dictionaries with content to write to the worksheet
    :param sheet_name: name of the odmlib worksheet to create to write variable content
    :param workbook: workbook object to add the worksheet to
    :param header_format: format for the header_row used when writing the column headers
    """
    worksheet = workbook.add_worksheet(sheet_name)
    write_header_row(worksheet, header_format)
    for r, row in enumerate(table):
        for c, col_name in enumerate(header):
            worksheet.write(r+1, c, row[col_name])


def process_map_sheet(sheet, domain, def_workbook, header_format):
    """
    process the content in the SDTM mapping spreadsheet to generate the define-xml metadata worksheet
    :param sheet: SDTM mapping spreadsheet worksheet to parse
    :param domain: 2-letter domain abbreviation used to identify the domain for a variable
    :param def_workbook: obmlib define-xml workbook to add variables worksheet to
    :param header_format: format for worksheet column headers
    """
    rows = []
    # adjust sheet.max_column to 60 if hidden columns are not counted
    for col_num, col in enumerate(sheet.iter_cols(min_col=2,max_col=sheet.max_column, min_row=1, max_row=10, values_only=True)):
        row_dict = {key: "" for key in header}
        if not col[0] or "CODELIST" in col[0]:
            continue
        for row_num, cell in enumerate(col):
            formatter = eval(format_class[row_num] + "()")
            formatter.format_content(cell, row_dict, domain, col_num)
        rows.append(row_dict)
    create_define_sheet(rows, sheet.title, def_workbook, header_format)


def set_cmd_line_args():
    """
    get the command-line arguments needed to convert the Excel input file into Define-XML
    :return: return the argparse object with the command-line parameters
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input_file", help="path and file name of the SDTM map spreadsheet file",
                        required=False, dest="map_xls", default=excel_map_file)
    parser.add_argument("-o", "--output_file", help="path and file name of Define-XML v2.1 metadata spreadsheet file",
                        required=False, dest="define_xls", default=excel_define_file)
    args = parser.parse_args()
    return args


def main():
    """
    main driver application that processes the SDTM mapping spreadsheet and creates and odmlib variable worksheet
    """
    args = set_cmd_line_args()
    def_workbook = XLS.Workbook(args.define_xls, {"strings_to_numbers": False})
    header_format = def_workbook.add_format({"bold": True, "bg_color": "#CCFFFF", "border": True, "border_color": "black"})
    map_workbook = load_workbook(filename=args.map_xls, read_only=False, data_only=True)
    for sheet in map_workbook.worksheets:
        if sheet.title not in worksheet_skip:
            print(f"processing {sheet.title}...")
            domain = sheet.title.split()
            process_map_sheet(sheet, domain[0], def_workbook, header_format)
    def_workbook.close()


if __name__ == '__main__':
    main()
