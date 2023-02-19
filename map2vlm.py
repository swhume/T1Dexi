from openpyxl import load_workbook
import xlsxwriter as XLS
import os.path
import json
import argparse

"""
map2vlm.py generates content for the ValueLevel, WhereClauses, and CodeLists worksheets in the T1Dexi odmlib
metadata spreadsheet. The output spreadsheet produced includes the previously mentioned worksheets and this content
can be pasted into the corresponding metadata worksheets. Content is pulled from the SDTM mapping spreadsheet, but the
mapping spreadsheet doesn't contain all the necessary information in a way that can be processed in a straightforward 
manner. The content is found in the CODELIST - VARIABLE NAME columns after the variable name columns.
The codelists dictionary listed below adds information needed to work with content from the mapping spreadsheet. 
If there are changes to the study that impact value level metadata this dictionary may need to
be updated.

Example Cmd-line (optional args):
    python map2vlm -i ./path/to/mapping_spec.xlsx -o ./path/to/define-worksheets.xlsx
"""

worksheet_skip = ["T1Dexi SDTM Summary", "T1Dexi Tables", "Domains", "Sheet1", "FALB"]
excel_map_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'SDTM-mapping-spec-27Jan2023.xlsx')
subset_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'cl_subsets.json')
excel_define_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'vlm-test.xlsx')
vlm_file = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'vlm.json')

# worksheet headers
cl_header = ["OID", "Name", "NCI Codelist Code", "Data Type", "Order", "Term", "NCI Term Code", "Decoded Value",
          "Comment", "IsNonStandard", "StandardOID"]
wc_header = ["OID", "Dataset", "Variable", "Comparator", "Value", "Comment"]
vlm_header = ["OID", "Order", "Dataset", "Variable", "ItemOID", "Where Clause", "Data Type", "Length",
              "Significant Digits", "Format", "Mandatory", "Codelist", "Origin Type", "Origin Source", "Pages",
              "Method", "Predecessor", "Comment"]


def write_subset_file(codelists):
    """
    write the codelist subsets dictionary to file as JSON
    """
    with open(subset_file, 'w') as file_out:
        json.dump(codelists, file_out)


def process_map_sheet(sheet, domain, codelists):
    """
    process the content in the SDTM mapping spreadsheet for content to use in the define-xml metadata worksheets
    :param sheet: SDTM mapping spreadsheet worksheet to parse
    :param domain: 2-letter domain abbreviation used to identify the domain for a variable
    """
    # assumes not more than 6 VLM (max_row=17)
    for col_num, col in enumerate(sheet.iter_cols(min_col=2,max_col=sheet.max_column, min_row=1, max_row=17, values_only=True)):
        # row_dict = {key: "" for key in header}
        # TODO refactor nested ifs
        if col and "CODELIST" in str(col[0]):
            print(f"processing codelist {col[0]}...")
            codelist, var_name = col[0].split(" - ")
            # only process subsets related to VLM
            variable_key = domain + "." + var_name
            if variable_key in codelists:
                subset_codes = []
                for row_num, cell in enumerate(col):
                    if row_num > 9 and cell:
                        subset_codes.append(cell)
                        # if more than one subset then have VLM and must lookup testcds
                print(f"found {len(subset_codes)} subsets for {var_name} in domain {domain}")
                codelists[variable_key]["domain"] = domain
                if subset_codes:
                    codelists[variable_key]["subset_terms"] = subset_codes


def process_codelists(codelists):
    """
    process the codelists dictionary that has the SDTM mapping spreadsheet content added to create codelist XLS rows
    :return: list of dictionaries with codelist worksheet rows
    """
    rows = []
    for key, cl in codelists.items():
        domain, variable = key.split(".")
        for idx, data_type in enumerate(cl["type"]):
            row = {key: "" for key in cl_header}
            if cl["IsNonStandard"][idx] == "No":
                continue
            try:
                wc_value = generate_wc_name(cl["whereclause"][idx][0]["value"][0])
            except IndexError as err:
                print(f"index error: index {idx} in {key}")
            row["OID"] = "CL." + domain + "." + variable + "." + wc_value
            row["Name"] = "Codelist for " + domain + " " + variable + " where " + wc_value
            if len(cl["IsNonStandard"][idx]) > 3:
                row["NCI Codelist Code"] = cl["IsNonStandard"][idx]
            else:
                row["NCI Codelist Code"] = ""
            row["Data Type"] = cl["type"][idx]
            order_num = 1
            if data_type == "text" and "subset_terms" in cl:
                try:
                    for term in cl["subset_terms"][idx].split(", "):
                        row["Order"] = order_num
                        term_row = row.copy()
                        term_row["Term"] = term
                        term_row["NCI Term Code"] = ""
                        term_row["Decoded Value"] = ""
                        term_row["Comment"] = ""
                        if cl["IsNonStandard"][idx] == "Yes":
                            term_row["IsNonStandard"] = "Yes"
                            term_row["StandardOID"] = ""
                        else:
                            term_row["IsNonStandard"] = ""
                            term_row["StandardOID"] = "STD.2"
                        rows.append(term_row)
                        order_num += 1
                except Exception as e:
                    print(f"error in term {idx} with datatype {data_type}. Error message: {str(e)}")
    return rows


def find_cl_max_length(codelist):
    """
    find and return the length of the longest term in the codelist
    :param codelist: codelist of terms
    :return: integer length of the longest codelist item
    """
    max_length = 0
    for term in codelist.split(", "):
        if len(term) > max_length:
            max_length = len(term)
    return max_length


def find_number_length(domain, variable):
    """
    placeholder to determine the length of numeric fields assigned as value level metadata
    :param domain: the domain / dataset in which the variable exists
    :param variable: the name of the variable
    :return: integer indicating number length - UPDATE THIS DEFAULT VALUE TO GET ACTUAL LENGTH
    """
    return 3


def find_number_sigdigits(domain, variable):
    """
    placeholder to determine the number of significant digits for numeric fields assigned as value level metadata
    :param domain: the domain / dataset in which the variable exists
    :param variable: the name of the variable
    :return: integer indicating number of significant digits - UPDATE THIS DEFAULT TO GET ACTUAL SIGNIFICANT DIGITS
    """
    return 2


def process_where_clauses(codelists):
    """
    process the codelist dictionary with SDTM mapping content to generate the Where Clause metadata XLS rows
    :return: list of dictionaries with where clause worksheet rows
    """
    rows = []
    for key, cl in codelists.items():
        domain, variable = key.split(".")
        # for idx, wc in enumerate(cl["whereclause"]):
        for idx, wc in enumerate(cl["whereclause"]):
            first_variable = ""
            for wc_test in wc:
                row = {key: "" for key in wc_header}
                row["Dataset"] = domain
                check_values = []
                row["Variable"] = wc_test["variable"]
                if not first_variable:
                    first_variable = wc_test["variable"]
                row["Comparator"] = wc_test["comparator"]
                for cv in wc_test["value"]:
                    check_values.append(cv)
                row["Value"] = "|".join(check_values)
                # wc_value = generate_wc_name(check_values[0])
                # if multiple checks are used in a WhereClause the wc_value won't be unique so we add a number
                row["OID"] = "WC." + domain + "." + variable + "." + first_variable + "." + str(idx+1)
                row["Comment"] = ""
                rows.append(row)
    return rows


def process_vlm(codelists):
    """
    process the codelist dictionary with SDTM mapping content to generate the Value Level metadata XLS rows
    :return: list of dictionaries with value level metadata worksheet rows
    """
    rows = []
    for key, cl in codelists.items():
        domain, variable = key.split(".")
        for idx, wc in enumerate(cl["whereclause"]):
            row = {key: "" for key in wc_header}
            wc_variable = wc[0]["variable"]
            wc_value = generate_wc_name(wc[0]["value"][0])
            # if multiple checks are used in a WhereClause the wc_value won't be unique so we add a number
            row["Where Clause"] = "WC." + domain + "." + variable + "." + wc_variable + "." + str(idx+1)
            #TODO how do the VLM variables get added? May need to generate those variables here
            row["ItemOID"] = "IT." + domain + "." + variable + "." + wc_value
            row["OID"] = "VL." + domain + "." + variable
            row["Dataset"] = domain
            row["Variable"] = variable
            row["Data Type"] = cl["type"][idx]
            if row["Data Type"] == "text":
                row["Codelist"] = "CL." + domain + "." + variable + "." + wc_value
                row["Length"] = find_cl_max_length(row["Codelist"])
                row["Significant Digits"] = ""
                row["Format"] = ""
            elif row["Data Type"] == "integer":
                row["Codelist"] = ""
                row["Length"] = find_number_length(domain, variable)
                row["Significant Digits"] = ""
                row["Format"] = ""
            else:
                row["Codelist"] = ""
                row["Length"] = find_number_length(domain, variable)
                row["Significant Digits"] = find_number_sigdigits(domain, variable)
                row["Format"] = str(find_number_length(domain, variable)) + "." + str(row["Significant Digits"])
            #TODO how determine mandatory for VLM?
            row["Mandatory"] = "No"
            row["Order"] = idx + 1
            row["Origin Type"] = ""
            row["Origin Source"] = ""
            row["Pages"] = ""
            row["Method"] = ""
            row["Predecessor"] = ""
            row["Comment"] = ""
            rows.append(row)
    return rows


def generate_wc_name(wc_name):
    wc_name_nodash = wc_name.replace("-", " ")
    wc_name_nodash = " ".join(wc_name_nodash.split())
    wc_name_value = wc_name_nodash.replace(" ", "-")
    return wc_name_value


def write_codelist_to_xls(worksheet, rows, header, row_nbr=0):
    """
    write code list, value level, or where clause rows to the odmlib worksheet
    :param worksheet: odmlib worksheet object to write to
    :param rows: a list of dictionaries (rows) to write to the worksheet
    :param row_nbr: integer row number that indicates which worksheet row to begin appending terms
    :return: integer row number that indicates the next row to begin appending content
    """
    for row in rows:
        row_nbr += 1
        for c, col_name in enumerate(header):
            worksheet.write(row_nbr, c, row[col_name])
    return row_nbr


def write_header_row(worksheet, header, header_format):
    """
    write the worksheet header column labels to a given odmlib worksheet
    write the header row for a given worksheet
    :param worksheet: the odmlib worksheet to write the column headers to
    :param header: the header for the worksheet to be generated
    :param header_format: the header format to indicate the style of the header columns
    """
    for c, header_label in enumerate(header):
        worksheet.write(0, c, header_label, header_format)


def set_cmd_line_args():
    """
    get the command-line arguments needed to convert the Excel input file into Define-XML
    :return: return the argparse object with the command-line parameters
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input_file", help="path and file name of the SDTM map spreadsheet file",
                        required=False, dest="map_xls", default=excel_map_file)
    parser.add_argument("-o", "--output_file", help="path and file name of Define-XML v2.1 metadata spreadsheet file",
                        required=False, dest="define_xls", default=excel_define_file)
    args = parser.parse_args()
    return args


def main():
    """
    main driver for creating the odmlib worksheet for ValueLists, WhereClauses, and CodeLists
    """
    args = set_cmd_line_args()
    def_workbook = XLS.Workbook(args.define_xls, {"strings_to_numbers": False})
    header_format = def_workbook.add_format({"bold": True, "bg_color": "#CCFFFF", "border": True, "border_color": "black"})
    #TODO re-establish automatically generating the vlm.json file
    map_workbook = load_workbook(filename=args.map_xls, read_only=False, data_only=True)
    codelists = json.load(open(vlm_file))

    for sheet in map_workbook.worksheets:
        if sheet.title not in worksheet_skip:
            print(f"processing {sheet.title}...")
            domain = sheet.title.split()
            process_map_sheet(sheet, domain[0], codelists)
    rows = process_codelists(codelists)
    write_subset_file(codelists)

    # create codelist subset worksheet
    worksheet = def_workbook.add_worksheet("codelists")
    write_header_row(worksheet, cl_header, header_format)
    write_codelist_to_xls(worksheet, rows, cl_header)

    # create where clause worksheet
    rows = process_where_clauses(codelists)
    wc_worksheet = def_workbook.add_worksheet("whereclauses")
    write_header_row(wc_worksheet, wc_header, header_format)
    write_codelist_to_xls(wc_worksheet, rows, wc_header)

    # create VLM
    rows = process_vlm(codelists)
    vlm_worksheet = def_workbook.add_worksheet("valuelevel")
    write_header_row(vlm_worksheet, vlm_header, header_format)
    write_codelist_to_xls(vlm_worksheet, rows, vlm_header)

    def_workbook.close()


if __name__ == '__main__':
    main()
